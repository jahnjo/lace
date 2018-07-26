import bs4 
from openpyxl import load_workbook
from urllib.request import urlopen as uReq
from bs4 import BeautifulSoup as soup
import sys
import random
import string
import datetime

urls = ['https://www.lacelab.com/collections/luxury-leather-laces/products/black-luxury-leather-laces-gold-plated',
'https://www.lacelab.com/collections/3m-reflective-rope-laces/products/black-3m-reflective-rope-laces',
'https://www.ropelacesupply.com/collections/all-shoe-laces/products/g-o-a-t-flat-shoe-laces',
'https://www.ropelacesupply.com/collections/premium-leather-shoe-laces/products/gold-leather-laces-1',
'https://www.ropelacesupply.com/collections/3m-reflective-shoe-laces/products/round-3m-black',
'https://www.lacelab.com/collections/thin-rope-laces/products/white-roshe-style-laces',
'https://www.lacelab.com/collections/oval-sb-foamposites-laces/products/blue-oval-sb-laces']

#Containers
laceData = ''
excelFile = 'lace.xlsx'

def createTitleBars(currentSheet):
    currentSheet['A1'] = 'Name'
    currentSheet['B1'] = 'Price'
    currentSheet['C1'] = 'Inventory'

def createWorkBook(excelFile):
    wb = load_workbook(excelFile)
    return wb

def excelInit(wb):   
    now = datetime.datetime.now()
    currentDate = now.strftime("%m-%d-%Y")
    if currentDate not in wb.sheetnames:
        currentSheet = wb.create_sheet(currentDate)
    else:
        currentSheet = wb.active
    return currentSheet

def populateSheet(currentSheet):
    for x in range(0,len(nameList)):
        currentSheet['A{}'.format(x+2)] = nameList[x]
        currentSheet['B{}'.format(x+2)] = priceList[x]
        currentSheet['C{}'.format(x+2)] = quantList[x]
    

#Used in order to find out how many different sizes/types of laces are on the page
sampleString = "inventory_quantity"
#Containers to hold all the information
rawList = []
nameList = []
priceList =[]
quantList = []

def randWorksheet():
    rand = ''.join(random.choice(string.ascii_lowercase + string.ascii_uppercase) for _ in range(5))
    return rand

def urlRequest(currentURL):
    #Connecting, Downloading, Closing webpage
    uClient = uReq(currentURL)
    pageHTML = uClient.read()
    uClient.close()
    return pageHTML

def parseHTML(page):
    #pageSoup holds all the HTML within bs4 object
    pageSoup = soup(page, "html.parser")
    return pageSoup

def getTag(rawHTML):
    #Holds all the HTML within <script type="text/javascript"> tags
    scriptTags = rawHTML.findAll("script",{"type":"text/javascript"})
    return scriptTags

def getLaceData(rawScript):
    currentTag = []
    global laceData
    #Going through each <script> tag and finding the right one with the lace data
    for x in range(0,len(rawScript)):
        #Casting bs4 object to string to use find() function
        currentTag.append(str(rawScript[x]))

        #lacelab.com algorithm
        if currentTag[x].find("new Shopify") != -1:
            #print("found 1")
            laceData = currentTag[x]

        #ropelacesupply.com algorithm    
        elif currentTag[x].find("var json_product") != -1:
            #print("found 2")
            laceData = currentTag[x]

def getBeginning():
    global laceData
    #Aligning the parse to accurately iterate through each item
    laceData = laceData[laceData.find(':[{"id"'):]

def getOutOfStock(x):
    global laceData
    global nameList
    global priceList
    global quantList
    #Checking if the item is sold out
    if laceData.count(sampleString) == 0:
        name = urls[x][urls[x].rfind("/") + 1 : ]
        print(name + " sold out")
        nameList[x] = name
        priceList[x] = "SOLD OUT"
        quantList[x] = "SOLD OUT"
        return True
    else:
        return False

def getData():
    global laceData
    global sampleString
    global rawList
    global nameList
    global priceList
    global quantList
    #Parsing through each item and putting name, price, quantity into their respective container.
    for x in range(0,laceData.count(sampleString)):
        rawList.append(laceData[laceData.find('{"id"') + 2 : laceData.find("</script>")]) 
        laceData = laceData[laceData.find('}')+1:]

        if rawList[x].find("\"name\""):
            nameList.append(rawList[x][rawList[x].find("\"name\"") + 8 : rawList[x].find('public_title') - 5 ])

        if rawList[x].find("\"price\""):
            temp = rawList[x][rawList[x].find("\"price\"") + 8 : rawList[x].find('weight') - 2 ]
            priceList.append(str(format(int(temp)/100, '.02f')))

        if rawList[x].find("\"inventory_quantity\""):
            quantList.append(rawList[x][rawList[x].find("\"inventory_quantity\"") + 21 : rawList[x].find('inventory_management') - 2 ]) 

def clearData():
    global laceData
    global rawList
    laceData = ''
    rawList = []

#Main
for x in range(0,len(urls)):
    dlContent = urlRequest(urls[x])
    HTMLsoup = parseHTML(dlContent)
    rawScriptData = getTag(HTMLsoup)
    getLaceData(rawScriptData)
    getBeginning()
    if getOutOfStock(x) is True:
        continue
    getData()
    clearData()

wb = createWorkBook(excelFile)
currentSheet = excelInit(wb)

createTitleBars(currentSheet)
populateSheet(currentSheet)

'''for x in range(0,len(nameList)):
    print(nameList[x])
    print(priceList[x]) 
    print(quantList[x])
    print()'''

wb.save(excelFile)















